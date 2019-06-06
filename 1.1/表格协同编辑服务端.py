import os
import sys
import socket
import random
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from multiprocessing.managers import BaseManager
from subprocess import Popen
from threading import Thread
from openpyxl import load_workbook,Workbook
from time import sleep


__author__ = 'jinke18'

HOST_NAME = socket.gethostname()
MANAGER_DOMAIN = socket.gethostbyname(HOST_NAME)
MANAGER_PORT = 27131
MANAGER_AUTH_KEY = b'1234'

#随机生成颜色，颜色作为客户端的唯一标识
COLOR= '#'
for i in range(6):
    COLOR+=random.choice(['0','1','2','3','4','5','6','7','8','9','A','B','C','D','E','F'])



INFO = {'name':HOST_NAME,'color':COLOR}

_idx = 1
class Cell(tk.Entry):
    def __init__(self,parent,row, column, text):#row,column从1开始
        self.row=row
        self.column=column
        self.var = tk.StringVar()
        self.var.trace("w", self.resetWidth)
        self.text = '' if not text else text
        super().__init__(parent,textvariable=self.var,
                       highlightbackground='black',
                       highlightcolor = INFO['color'],
                       highlightthickness=1,
                       takefocus=True)
        
        self.var.set(self.text)
        self.bind("<FocusIn>",self.onFocusIn)
        self.bind("<FocusOut>",self.onFocusOut)
    def put(self,**kwargs):
        self.grid(row=self.row-1, column=self.column-1, sticky=tk.NE+tk.SW, **kwargs)
    def read(self):
        text = self.var.get()
        if not text:
            return None
        else:
            return text
    def write(self,text):
        text = '' if not text else text
        self.var.set(text)
    def resetWidth(self,*args):
        '''单元格宽度自适应'''
        self.text = self.var.get()
        width=0
        for s in self.text:
            if '0'<s<'9'or s == ' ':
                width+=1
            elif 'a'<s<'z' or 'A'<s<'Z':
                width+=1.25
            else:
                width+=1.66
        self.width=int(width+1)
        self['width']=self.width
    def resetAll(self,text,attr):
        if attr['highlightbackground'] !=INFO['color']:
            self.write(text)
            print('write')
            if attr['highlightbackground'] != self['highlightbackground']:
                self.configure(attr)
    def onFocusIn(self,*args):
        sheetdata.reset((self.row,self.column),highlightbackground=INFO['color'],state='disabled')
    def onFocusOut(self,*args):
        print(self.var.get())
        sheetdata.reset((self.row,self.column),text=self.read(),highlightbackground='black',state='normal')
        print('FocusOut,resetdata')
class sheetFrame(tk.Frame):
    def __init__(self, root):
        tk.Frame.__init__(self, root)
        self.canvas = tk.Canvas(root, borderwidth=0)
        self.frame = tk.Frame(self.canvas)
        
        self.vsb = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.vsb.pack(side="right", fill="y")
        
        self.hsb=tk.Scrollbar(root,orient='horizontal',command=self.canvas.xview)
        self.hsb.pack(side='bottom', fill='x')

        self.canvas.configure(yscrollcommand=self.vsb.set,xscrollcommand=self.hsb.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((0,0), window=self.frame, anchor="nw", tags="sheetFrame")
        self.frame.bind("<Configure>", self.onFrameConfigure)
        self.canvas.bind_all('<MouseWheel>',self.onMouseWheel)
        self.cells={}
    def onMouseWheel(self,event):
        self.canvas.yview_scroll(-1*(event.delta//120),'units')
    def setcell(self,place,text):
        self.cells[place]=Cell(self.frame,*place,text)
        self.putcell(place)
    def putcell(self,place,**kwargs):
        self.cells[place].put(**kwargs)
    def onFrameConfigure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

class dataManager(BaseManager):
    pass

dataManager.register('sheetdata')

class DataClient():
    def __init__(self, domain, port, auth_key):
        self.domain = domain
        self.port = port
        self.auth_key = auth_key
        self.datamanager = dataManager(address=(self.domain, self.port), authkey=self.auth_key)
        self.datamanager.connect()
        print('connected to server')
    def get_sheetdata(self):
        self.sheetdata=self.datamanager.sheetdata()
        return self.sheetdata
def auto_reset():
    global _idx
    while 1:
        place = sheetdata.get_resetplace(_idx)
        if place:
            celldata = sheetdata.get_datadict(place)
            sheetframe.cells[place].resetAll(celldata['text'],celldata['attr'])
            _idx+=1
        else:
            sleep(2)




class dataSManager(BaseManager):
    pass

dataSManager.register('sheetdata', callable=lambda: _sheetdata)


def transcol(col):
    if isinstance(col,int) and col>=1:
        colname=""
        while 1:
            col-=1
            colname = chr(65+col%26) + colname
            col = col//26
            if not col:
                break
        return colname
    else:
        colidx=0
        while col:
            colidx*=26
            colidx += (ord(col[0])-64)
            col = col[1:]
        return colidx

def trans(arg):
    if isinstance(arg,str):
        for idx in range(len(arg)):
            if '0'<arg[idx]<'9':
                break
        col = arg[:idx]
        row = arg[idx:]
        return int(row),transcol(col)
    else:
        return transcol(arg[1])+str(arg[0])

def savedata(path):
    global wb

    ws = wb.active
    rows = ws.max_row
    columns = ws.max_column
    for row in range(1,1+rows):
        for column in range(1,1+columns):
            value = _sheetdata.datadict[(row,column)]['text']
            try:
                ws[trans((row,column))].value=value
            except AttributeError:
                pass
    if os.path.exists(path):
        path=path[:-5]+'_'+'.xlsx'
    try:
        wb.save(path)
    except:
        messagebox.showerror('错误','保持失败')
    return path
def getdata(xlspath):
    global wb
    datadict={}
    wb = load_workbook(xlspath)
    ws = wb.active
    rows = ws.max_row
    columns = ws.max_column
    print(rows,columns)
    for row in range(1,1+rows):
        for column in range(1,1+columns):

            value=ws[trans((row,column))].value
            datadict[(row,column)]={'text':value,'attr':{}}
    merged_cells=[]
    for cell in ws.merged_cells:
        merged_cells.append((cell.min_row,cell.min_col,cell.max_row-cell.min_row+1,cell.max_col-cell.min_col+1))
    return datadict,merged_cells
class SheetData:
    def __init__(self,xlspath):
        self.datadict,self.merged_cells=getdata(xlspath)
        print(len(self.datadict))
        self.resetdict=dict()
        self.idx=1
    def reset(self,place,text='',highlightbackground=None,state=None):
        if text!='':
            self.datadict[place]['text']=text
        if highlightbackground:
            self.datadict[place]['attr']['highlightbackground']=highlightbackground
        if state:
            self.datadict[place]['attr']['state']=state
        self.resetdict[self.idx]=place
        self.idx+=1
    def get_resetplace(self,idx):
        if idx in self.resetdict:
            return self.resetdict[idx]
        else:
            return {}
    def get_datadict(self,place=None):
        if place:
            if place=='get_merge':
                return self.datadict,self.merged_cells
            return self.datadict[place]
        else:
            return self.datadict
    def get_max_idx(self):
        if self.resetdict:
            return max(self.resetdict.keys())
        else:
            return 1


class DataServer():
    def __init__(self, domain, port, auth_key):
        self.domain = domain
        self.port = port
        self.auth_key = auth_key
    def start_manager_server(self):
        self.queue_manager = dataSManager(address=(self.domain, self.port), authkey=self.auth_key)
        self.server = self.queue_manager.get_server()
    def run(self):
        self.start_manager_server()
        self.server.serve_forever()
    def stop(self):
        self.server.shutdown()
        self.is_stop = 1
def startServer():
    server=DataServer(MANAGER_DOMAIN, MANAGER_PORT,MANAGER_AUTH_KEY)
    server.run()
def onClose():
    ans= messagebox.askyesnocancel('提示','注意：关闭该界面后将无法同步更改操作\n你有内容尚未保存，\n是否保存已更改内容？')
    #print(ans)
    if ans==True:
        root.destroy()
        onSave()
    elif ans==False:
        root.destroy()
        savedata(_path)
    else:
        pass
def onSave():
    newpath = savedata(_path)
    print(newpath)
    print('explorer /select,'+newpath)
    Popen('explorer /select,'+newpath)
def auto_save():
    while 1:
        sleep(300)
        savedata(_path)
def get_path():
    if sys.argv[1:]:
        path=sys.argv[1]
    else:
        path = filedialog.askopenfilename(title='导入表格文件').replace('/','\\')
    return path

if __name__ == '__main__':
    root=tk.Tk()
    root.withdraw()
    _path = get_path()
    print(_path)
    if _path =='':
        root.destroy()
    elif _path[-5:]!='.xlsx':
        messagebox.showerror('错误','只支持导入xlsx格式的表格')
        root.destroy()
    root.deiconify()
    try:
        _sheetdata = SheetData(_path)
    except PermissionError:
        root.withdraw()
        messagebox.showerror('错误','导入表格失败')
    #启动服务器线程
    serverthread=Thread(target=startServer,args=())
    serverthread.setDaemon(True)
    serverthread.start()
    #启动本地表格编辑主进程
    client=DataClient( MANAGER_DOMAIN,MANAGER_PORT, MANAGER_AUTH_KEY)
    sheetdata=client.get_sheetdata()
    _idx = sheetdata.get_max_idx()
    #print('_idx',_idx)
    #界面初始化
    root.title('表格协同编辑--服务器IP地址'+MANAGER_DOMAIN )
    root.protocol('WM_DELETE_WINDOW',onClose)
    root.geometry('1200x750')
    menubar = tk.Menu(root)
    menubar.add_command(label='保存文件',command=onSave)
    root['menu']=menubar
    sheetframe = sheetFrame(root)
    datadict,merged_cells=sheetdata.get_datadict('get_merge')
    for place in datadict.keys():
        sheetframe.setcell(place,datadict[place]['text'])
    for cell in merged_cells:
        print(cell)
        srow,scol,rowspan,colspan=cell
        for row in range(srow,srow+rowspan):
            for col in range(scol,scol+colspan):
                sheetframe.cells[(row,col)].grid_forget()
        sheetframe.cells[(srow,scol)].put(rowspan = rowspan,columnspan=colspan)
    #启动自动同步线程
    reset_thread=Thread(target=auto_reset,args=())
    reset_thread.setDaemon(True)
    reset_thread.start()
    #启动自动保存线程
    auto_save_thread=Thread(target=auto_save,args=())
    auto_save_thread.setDaemon(True)
    auto_save_thread.start()
    #打开界面
    root.mainloop()
